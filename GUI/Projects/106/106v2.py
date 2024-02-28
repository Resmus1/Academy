import tkinter as tk
from tkinter import filedialog
from tkinter.messagebox import showinfo
from openpyxl import *
from openpyxl.styles import Font, Alignment


def open_file_surnames():
    """
    Открывает файл таблицы, и возвращает список строк из него.
    """
    global surname_list
    sheet = open_exel()
    # Проходимся по каждой строке и выводим значения ячеек в список
    for row in sheet.iter_rows(values_only=True):
        if row[0] is not None:
            surname_list.append(row[0])

    showinfo(title="Информация", message="Файл с Фамилиями загружен")


def open_exel():
    """
    Открывает файл ексель и получает активный лист
    """
    # Укажите путь к вашему файлу Excel
    file_path = filedialog.askopenfilename()
    # Открываем файл Excel
    active_workbook = load_workbook(file_path)
    # Получаем активный лист
    sheet = active_workbook.active
    return sheet


def edit_exel(ws):
    """
    Редактирование пустого exel
    """
    # Запись заголовка
    ws.append(['ФИО', 'Дата'])
    # Устанавливаем шрифт для текста в указанных ячейках выравнивает их по центру
    cells_to_format_font = ['A1', 'B1']
    for cell in cells_to_format_font:
        ws[cell].font = Font(size=20, bold=True)
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
    # изменяем ширину колонки
    cells_to_format_size = ['A', 'B']
    for cell in cells_to_format_size:
        ws.column_dimensions[cell].width = 30
        ws.column_dimensions[cell].font = Font(size=14)
    return ws


def generate_list(sheet, list_data, ws):
    """
    Проходимся по каждой строке и выводим значения ячеек в словарь,
    а затем отправляем в список и записываем в ексель
    """
    for row in sheet.iter_rows(values_only=True):
        if row[0] in surname_list:
            i = 0
            list_data[row[0]] = []
            for data in row[1:]:
                i += 1
                if data is not None:
                    list_data[row[0]].append(str(i))
    # Переносит список в ексель
    for surname, data_list in list_data.items():
        data_str = [', '.join(data_list)]
        row_data = [surname] + data_str
        ws.append(row_data)
    return ws


def generation_file_list():
    """
    Создание нового файла ексель, запись сгенерированого списка и сохранение
    """
    sheet = open_exel()
    # Создание нового файла Excel
    wb = Workbook()
    # Активация активного листа
    ws = wb.active

    edit_exel(ws)

    generate_list(sheet, list_data, ws)

    # Сохранение файла
    file_name = '106.xlsx'
    wb.save(file_name)

    showinfo(title="Информация", message="Файл с 106-й создан")
    exit()


def open_info():
    showinfo(title="Информация", message="Информационное сообщение")


surname_list = []

list_data = {}

win = tk.Tk()
win.geometry(f"400x200+100+200")
win.title("Создание списка по 106-й")
icon = tk.PhotoImage(file='cleaning_icon.png')
win.iconphoto(False, icon)
win.resizable(False, False)

tk.Label(text="Создание списка по 106-й",
         font=('Arial', 20, 'italic', 'bold'),
         ).grid(row=0)

tk.Label(text="Выберете файл с Фамилиями которые нужно найти:",
         font=('Arial', 11, 'italic'),
         ).grid(row=1)

tk.Label(text="Выберете файл со 106-й:",
         font=('Arial', 11, 'italic'),
         ).grid(row=3)

tk.Button(win, text='Выберете Фамилии',
          command=open_file_surnames,
          font=('Arial', 10, 'bold'),

          ).grid(row=2, sticky='we')

tk.Button(win, text='Генерация',
          command=generation_file_list,
          font=('Arial', 10, 'bold'),
          ).grid(row=4, sticky='we')

win.mainloop()


# Левый краешек подредактировать, сделать отступ