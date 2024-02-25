from openpyxl import *

# Укажите путь к вашему файлу Excel
file_path = "surnames_list.xlsx"
# Открываем файл Excel
workbook = load_workbook(file_path)
# Получаем активный лист
sheet = workbook.active

# Создание нового файла Excel
wb = Workbook()
# Активация активного листа
ws = wb.active

surnames = ('Смирнов А.И.', 'Иванов В.П.', 'Кузнецов Г.М.', 'Попов Д.С.')
value31 = [i for i in range(1, 32)]

# Запись заголовка
ws.append(['ФИО', *value31])
# изменяем ширину колонки
ws.column_dimensions['A'].width = 20

# Проходимся по каждой строке и выводим значения ячеек
for row in sheet.iter_rows(values_only=True):
    if row[0] in surnames:
        ws.append(row)

# Сохранение файла
file_name = '106.xlsx'
wb.save(file_name)

# Закрываем файл Excel
workbook.close()
