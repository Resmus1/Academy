import tkinter as tk
from tkinter import filedialog
from tkinter.messagebox import showinfo
from openpyxl import *
from openpyxl.styles import Font, Alignment


def open_file_surnames():
    """
    Открывает активный лист, и возвращает список строк из него, преобразуя в список питон
    """
    global surname_list
    sheet = open_exel()
    for row in sheet.iter_rows(values_only=True):
        if row[0] is not None:
            surname_list.append(row[0])
    print(surname_list)
    showinfo(title="Информация", message="Файл с Фамилиями загружен")


def open_exel():
    """
    Открывает файл ексель и получает активный лист
    """
    file_path = filedialog.askopenfilename()
    active_workbook = load_workbook(file_path)
    sheet = active_workbook.active
    return sheet


def edit_exel(ws):
    """
    Редактирование пустого exel
    """
    # Запись заголовка
    ws.append(surname_list)
    # Устанавливаем шрифт для текста в указанных ячейках выравнивает их по центру
    cells_to_format_font = ['A1', 'B1']
    for cell in cells_to_format_font:
        ws[cell].font = Font(size=20, bold=True)
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
    # изменяем ширину колонки
    cells_to_format_size = ['A', 'B']
    for cell in cells_to_format_size:
        ws.column_dimensions[cell].width = 30
        ws.column_dimensions[cell].font = Font(size=14)
    return ws


def generation_file_list():
    """
    Создание нового файла ексель, запись сгенерированого списка и сохранение
    """
    sheet = open_exel()
    # Создание нового файла Excel
    wb = Workbook()
    # Активация активного листа
    ws = wb.active

    edit_exel(ws)

    # Сохранение файла
    file_name = '106.xlsx'
    wb.save(file_name)

    showinfo(title="Информация", message="Файл с 106-й создан")
    exit()


surname_list = []

list_data = {}

win = tk.Tk()
win.geometry(f"400x200+100+200")
win.title("Создание бирок")
icon = tk.PhotoImage(file='img/tag.png')
win.iconphoto(False, icon)
win.resizable(False, False)

tk.Label(win, text="Создание бирок по списку фамилий",
         font=('Arial', 11, 'italic'),
         ).grid()

tk.Button(win, text='Выберете Фамилии',
          command=open_file_surnames,
          font=('Arial', 10, 'bold'),
          ).grid(row=2, sticky='we')

tk.Button(win, text='Генерация',
          command=generation_file_list,
          font=('Arial', 10, 'bold'),
          ).grid(row=4, sticky='we')

win.mainloop()
