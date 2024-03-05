from openpyxl import *
from openpyxl.styles import Font, Alignment, Border, Side


def create_table():
    """
    Создание консольной таблицы для переноса в ексель.
    """
    global table
    # Открываем файл Excel
    workbook = load_workbook("list_on_print.xlsx")
    # Получаем активный лист
    sheet = workbook.active
    # Проходим по строкам выводим значения в список
    for row in sheet.iter_rows(values_only=True):
        if row[0] is not None:
            table.append(row[0])
    # проходим по списку и разбиваем на по 3
    table = [table[i:i + 3] for i in range(0, len(table), 3)]


def style_table():
    """
    Задает Высоту и ширину ячеек.
    """
    cells_font = ['A', 'B', 'C']
    for cell in cells_font:
        ws.column_dimensions[cell].width = 30
    for i in range(11):
        ws.row_dimensions[i].height = 60


def style_cells():
    """
    Рисование границ ячеек и выравнивание текста
    """
    letter_list = ['A', 'B', 'C']
    thins = Side(border_style="thin", color="000000")
    for letter in letter_list:
        # Выравнивание текста задание размера
        col = ws.column_dimensions[f'{letter}']
        col.alignment = Alignment(horizontal="center", vertical="center")
        col.font = Font(size=14)
        # Рисование границ ячеек
        col.border = Border(top=thins, bottom=thins, left=thins, right=thins)


wb = Workbook()
# Активация активного листа
ws = wb.active

table = []

create_table()
style_table()

for row in table:
    ws.append(row)

style_cells()

# Сохранение файла
wb.save('test.xlsx')
