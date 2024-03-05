from openpyxl import *
from openpyxl.styles import Font, Alignment

# Укажите путь к вашему файлу Excel
file_path = "surnames_list.xlsx"
# Открываем файл Excel
workbook = load_workbook(file_path)
# Получаем активный лист
sheet = workbook.active

# Создание нового файла Excel
wb = Workbook()
# Активация активного листа
ws = wb.active

surnames = ('Смирнов А.И.', 'Иванов В.П.', 'Кузнецов Г.М.', 'Попов Д.С.')
value31 = [i for i in range(1, 32)]

# Запись заголовка
ws.append(['ФИО', 'Дата'])

# Устанавливаем шрифт для текста в указанных ячейках выравнивает их по центру
cells_to_format_font = ['A1', 'B1']
for cell in cells_to_format_font:
    ws[cell].font = Font(size=20, bold=True)
    ws[cell].alignment = Alignment(horizontal='center', vertical='center')

# изменяем ширину колонки
cells_to_format_size = ['A', 'B']
for cell in cells_to_format_size:
    ws.column_dimensions[cell].width = 30
    ws.column_dimensions[cell].font = Font(size=14)

list_data = {}

# Проходимся по каждой строке и выводим значения ячеек в список
for row in sheet.iter_rows(values_only=True):
    if row[0] in surnames:
        i = 0
        list_data[row[0]] = []
        for data in row[1:]:
            i += 1
            if data is not None:
                list_data[row[0]].append(str(i))

# Переносит список в ексель
for surname, data_list in list_data.items():
    data_str = [', '.join(data_list)]
    row_data = [surname] + data_str
    ws.append(row_data)

# Сохранение файла
file_name = '106.xlsx'
wb.save(file_name)
