from openpyxl import *
from openpyxl.styles import Font, Alignment, Border, Side

data = {}


def open_surname_list():
    """
    Открывает файл таблицы с фамилиями, и возвращает словарь с ключами из них.
    """
    #  Открытие списка Фамилий на поиск
    surname_excel = load_workbook('/home/deck/Desktop/Squad/Общий список.xlsx')
    sheet = surname_excel.active
    # Удаляет пробелы по бокам и создает словарь с ключами
    for row in sheet.iter_rows(values_only=True):
        if row[0] is not None:
            data[row[0].strip()] = []


def generate_dict_106():
    """
    Проходится по базе и добавляет в словарь data значения к ключам.
    """
    # Открытие базы с графиками
    base_chart = load_workbook('/home/deck/Desktop/Squad/График март upd2.xlsx')
    sheet = base_chart.active
    # Проходится по всему списку в базе
    for row in sheet.iter_rows(values_only=True):
        # Проходится по столбцу, находит строки соответствующие условию
        if isinstance(row[1], str) and len(row[1]) < 22:
            # Обрезает слово для проверки оставляя одну фамилию и присваивает значение
            value = row[1].replace('.', '').strip()[0:-3].lower()
            # Проходится по словарю сравнивая значения
            for surname in data.keys():
                # Так же обрезается до фамилии при совпадении идет далее
                if value == surname.replace('.', '').strip()[0:-3].lower():
                    # Проходится по совпавшей строке добавляет число при совпадении к ключу
                    day_num = 0
                    for day in row[2:]:
                        day_num += 1
                        if day in ('х', 'Х', 'дх', 'дХ'):
                            data[surname].append(day_num)


def create_106_file():
    """
    Создание документа.
    """
    book, act_list = create_list()
    # Проходится по словарю и добавляет значения в активный лист
    i = 1
    for name, numbers in data.items():
        i += 1
        if i < 43:
            act_list[f'A{i}'] = i - 1
            act_list[f'B{i}'] = name
            act_list[f'C{i}'] = str(numbers)[1:-1]
        else:
            act_list[f'E{i - 41}'] = i - 1
            act_list[f'F{i - 41}'] = name
            act_list[f'G{i - 41}'] = str(numbers)[1:-1]
    # Сохранение файла
    book.save('/home/deck/Desktop/Squad/106.xlsx')


def create_list():
    """
    Создание и редактирование документа excel
    """
    # Создание нового файла Excel
    new_workbook = Workbook()
    # Активация активного листа
    active_list = new_workbook.active

    # Толщина и цвет рамки
    thins = Side(border_style='thin', color='000000')

    # Запись заголовка
    active_list.append(['№', 'ФИО', 'Дата', '', '№', 'ФИО', 'Дата'])
    # Заголовок, выравнивание и установка шрифта
    heading = ['A1', 'B1', 'C1', 'E1', 'F1', 'G1']
    for cell in heading:
        active_list[cell].font = Font(size=15, bold=True)
        active_list[cell].alignment = Alignment(horizontal='center', vertical='center')
        active_list[cell].border = Border(top=thins, bottom=thins, left=thins, right=thins)

    # изменяем ширину колонки установка позиции текста
    column = active_list.column_dimensions
    column['A'].width = 4
    column['A'].alignment = Alignment(horizontal='center')
    column['B'].width = 24
    column['C'].width = 15
    column['D'].width = 3
    column['E'].width = 4
    column['E'].alignment = Alignment(horizontal='center')
    column['F'].width = 24
    column['G'].width = 15
    # Шрифт для ячеек
    column_table = ['A', 'B', 'C', 'E', 'F', 'G']
    for col in column_table:
        column[col].font = Font(size=12)
        # Рисование границ ячеек
        column[col].border = Border(top=thins, bottom=thins, left=thins, right=thins)

    return new_workbook, active_list


open_surname_list()
generate_dict_106()
create_106_file()
